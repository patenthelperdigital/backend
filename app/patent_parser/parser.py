import re
from io import BytesIO

import aiofiles
import openpyxl
from fastapi import HTTPException
from starlette.responses import FileResponse


async def create_upload_file(file) -> FileResponse:
    """
     Асинхронная функция для обработки загруженного файла Excel с данными о патентах.

     Args:
         file (UploadFile): Загруженный файл Excel.

     Raises:
         HTTPException: Если формат файла не поддерживается.

     Returns:
         FileResponse: Ответ с созданным файлом Excel, содержащим обновленные данные о патентах.

     Эта функция выполняет следующие действия:
     1. Проверяет формат загруженного файла. Если формат не соответствует формату Excel, выбрасывается исключение HTTPException.
     2. Читает содержимое загруженного файла.
     3. Загружает рабочую книгу Excel из содержимого файла.
     4. Обрабатывает данные из активного листа рабочей книги.
     5. Создает новую рабочую книгу Excel и записывает в нее обработанные данные.
     6. Сохраняет новую рабочую книгу в файловый поток.
     7. Асинхронно записывает содержимое файлового потока в новый файл на диске.
     8. Возвращает FileResponse с созданным файлом Excel.
     """
    if file.content_type != 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        raise HTTPException(status_code=400, detail="Данный формат файла не поддерживается.")

    try:
        content = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        results = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            registration_number = row[0]
            patent_holders = row[6]
            if patent_holders:
                holder_list = [re.sub(r'\s*\(RU\)$', '', holder.strip() + ')') for holder in patent_holders.split(')')[:-1]]

                holder_count = len(holder_list)
                results.append({
                    "registration_number": registration_number,
                    "patent_holders": holder_list,
                    "holder_count": holder_count
                })

        output_workbook = openpyxl.Workbook()
        output_sheet = output_workbook.active
        output_sheet.append(["registration_number", "patent_holders", "holder_count"])

        for result in results:
            output_sheet.append([
                result["registration_number"],
                ", ".join(result["patent_holders"]),
                result["holder_count"]
            ])

        output_path = "updated_patents.xlsx"

        output_stream = BytesIO()
        output_workbook.save(output_stream)
        output_stream.seek(0)

        async with aiofiles.open(output_path, 'wb') as f:
            await f.write(output_stream.getvalue())

        output_stream.seek(0)

        return dict(status="OK")
    except Exception as e:
        return dict(status="Error", msg=str(e))
